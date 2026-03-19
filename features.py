# features.py — Extra Features (Fixed Version)

import io, random
from datetime import datetime, timedelta

# ── QR CODE ───────────────────────────────────────────────────────────────────
def generate_qr_code(machine_id, base_url="http://127.0.0.1:5000"):
    try:
        import qrcode
        import base64
        url = f"{base_url}/search?machine_id={machine_id}"
        qr  = qrcode.QRCode(version=1, box_size=6, border=2)
        qr.add_data(url)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        return base64.b64encode(buf.read()).decode('utf-8')
    except Exception:
        return None

# ── EXCEL EXPORT ──────────────────────────────────────────────────────────────
def generate_excel_report(predictions):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        wb  = Workbook()
        ws  = wb.active
        ws.title = "Machine Health Report"

        # Styles
        hdr_font  = Font(bold=True, name="Calibri", size=11)
        title_font= Font(bold=True, name="Calibri", size=13)
        center    = Alignment(horizontal='center', vertical='center')
        thin      = Side(style='thin', color="CCCCCC")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Title row
        ws.merge_cells('A1:L1')
        ws['A1'] = f"Industrial Predictive Maintenance Report — {datetime.now().strftime('%d %B %Y %H:%M')}"
        ws['A1'].font      = title_font
        ws['A1'].alignment = center
        ws.row_dimensions[1].height = 25

        # Headers
        headers = ["Machine ID","Name","Operator","Location","Health Score %",
                   "Failure Risk %","Status","RUL Days","Failure Type",
                   "Last Maintenance","Est. Cost (Rs.)","Anomaly"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font      = hdr_font
            cell.alignment = center
            cell.border    = border
        ws.row_dimensions[2].height = 18

        # Data rows
        for row, (mid, pred) in enumerate(predictions.items(), 3):
            vals = [
                mid,
                pred['machine_info']['name'],
                pred['machine_info']['operator'],
                pred['machine_info']['location'],
                pred['health_score'],
                pred['failure_probability'],
                pred['status'],
                pred['rul_days'],
                pred['failure_type'],
                pred['machine_info']['last_maintenance'],
                pred['cost_estimate']['total_estimated'],
                "YES" if pred['is_anomaly'] else "NO",
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = center
                cell.border    = border

        # Column widths
        widths = [12,25,18,12,14,14,12,12,18,18,18,10]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = w

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2['A1'] = "Summary Statistics"
        ws2['A1'].font = Font(bold=True, size=13)
        total      = len(predictions)
        critical   = sum(1 for p in predictions.values() if p['status']=='Critical')
        warning    = sum(1 for p in predictions.values() if p['status']=='Warning')
        healthy    = sum(1 for p in predictions.values() if p['status']=='Healthy')
        avg_health = round(sum(p['health_score'] for p in predictions.values())/total, 1)
        total_cost = sum(p['cost_estimate']['total_estimated'] for p in predictions.values())
        rows2 = [
            ("Total Machines", total),
            ("Critical", critical),
            ("Warning",  warning),
            ("Healthy",  healthy),
            ("Avg Health Score", f"{avg_health}%"),
            ("Total Est. Cost",  f"Rs.{total_cost:,}"),
            ("Generated", datetime.now().strftime('%d %B %Y %H:%M')),
        ]
        for r, (k, v) in enumerate(rows2, 3):
            ws2.cell(row=r, column=1, value=k).font  = Font(bold=True)
            ws2.cell(row=r, column=2, value=str(v))
        ws2.column_dimensions['A'].width = 22
        ws2.column_dimensions['B'].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    except Exception as e:
        # Return a minimal valid xlsx if something fails
        from openpyxl import Workbook
        wb  = Workbook()
        ws  = wb.active
        ws['A1'] = f"Error generating report: {str(e)}"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

# ── OEE CALCULATOR ────────────────────────────────────────────────────────────
def calculate_oee(machine_id, health_score, failure_prob):
    random.seed(hash(machine_id) % 500)
    availability = round(max(50, health_score * 0.95 + random.uniform(-5, 5)), 1)
    performance  = round(max(50, 100 - failure_prob * 0.6 + random.uniform(-3, 3)), 1)
    quality      = round(max(60, 100 - failure_prob * 0.4 + random.uniform(-2, 2)), 1)
    oee          = round((availability * performance * quality) / 10000, 1)
    return {
        "availability": min(100, availability),
        "performance":  min(100, performance),
        "quality":      min(100, quality),
        "oee":          min(100, oee),
        "rating": ("Excellent" if oee >= 85 else "Good" if oee >= 70
                   else "Average" if oee >= 55 else "Poor")
    }

# ── WORK ORDERS ───────────────────────────────────────────────────────────────
_work_orders = []
_wo_counter  = 100

def get_work_orders():
    return _work_orders

def create_work_order(machine_id, issue, priority, assigned_to, created_by):
    global _wo_counter
    _wo_counter += 1
    wo = {
        "id":          f"WO-{_wo_counter}",
        "machine_id":  machine_id,
        "issue":       issue,
        "priority":    priority,
        "assigned_to": assigned_to,
        "created_by":  created_by,
        "status":      "Open",
        "created_at":  datetime.now().strftime("%Y-%m-%d %H:%M"),
        "updated_at":  datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    _work_orders.insert(0, wo)
    return wo

def update_work_order_status(wo_id, status):
    for wo in _work_orders:
        if wo["id"] == wo_id:
            wo["status"]     = status
            wo["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            return True
    return False

def seed_work_orders():
    if not _work_orders:
        samples = [
            ("MCH-104", "High temperature warning",      "High",     "engineer1", "admin"),
            ("MCH-105", "Vibration anomaly detected",    "Critical", "engineer1", "admin"),
            ("MCH-102", "Scheduled bearing inspection",  "Medium",   "engineer1", "manager1"),
            ("MCH-101", "Routine lubrication check",     "Low",      "engineer1", "operator1"),
        ]
        for m, i, p, a, c in samples:
            create_work_order(m, i, p, a, c)

seed_work_orders()

# ── SPARE PARTS INVENTORY ─────────────────────────────────────────────────────
INVENTORY = [
    {"id":"INV-001","part":"Bearing Set",        "quantity":12,"min_qty":5, "unit_cost":3500,  "compatible":["MCH-101","MCH-102","MCH-104"],"status":"OK"},
    {"id":"INV-002","part":"Cooling Fan",         "quantity":3, "min_qty":4, "unit_cost":2200,  "compatible":["MCH-101","MCH-103"],          "status":"Low"},
    {"id":"INV-003","part":"Pressure Valve",      "quantity":8, "min_qty":3, "unit_cost":4500,  "compatible":["MCH-102","MCH-104","MCH-105"],"status":"OK"},
    {"id":"INV-004","part":"Drive Belt",          "quantity":15,"min_qty":5, "unit_cost":1200,  "compatible":["MCH-103"],                    "status":"OK"},
    {"id":"INV-005","part":"Lubrication Oil (L)", "quantity":45,"min_qty":20,"unit_cost":350,   "compatible":["MCH-101","MCH-102","MCH-103","MCH-104","MCH-105","MCH-106"],"status":"OK"},
    {"id":"INV-006","part":"Turbine Blade",       "quantity":2, "min_qty":3, "unit_cost":18000, "compatible":["MCH-106"],                    "status":"Low"},
    {"id":"INV-007","part":"Thermocouple Sensor", "quantity":6, "min_qty":4, "unit_cost":2800,  "compatible":["MCH-104","MCH-105"],          "status":"OK"},
    {"id":"INV-008","part":"Motor Coupling",      "quantity":0, "min_qty":2, "unit_cost":5500,  "compatible":["MCH-101","MCH-106"],          "status":"Out"},
]

def get_inventory():
    for item in INVENTORY:
        item["status"] = ("Out"  if item["quantity"] == 0
                          else "Low" if item["quantity"] <= item["min_qty"]
                          else "OK")
    return INVENTORY

# ── MAINTENANCE CHECKLIST ─────────────────────────────────────────────────────
def get_predictive_reorder(predictions):
    """
    Smart reorder alerts: cross-references ML predictions with inventory.
    If a machine is predicted to fail soon, flag required parts as URGENT.
    Returns list of reorder recommendations.
    """
    inventory = get_inventory()
    alerts = []
    for item in inventory:
        for mid in item.get('compatible', []):
            pred = predictions.get(mid, {})
            rul  = pred.get('rul_days', 999)
            fp   = pred.get('failure_probability', 0)
            status = item['status']
            # Urgent: machine failing soon + part low/out
            if (rul < 10 or fp > 60) and status in ('Low', 'Out'):
                alerts.append({
                    'part':     item['part'],
                    'inv_id':   item['id'],
                    'machine':  mid,
                    'status':   status,
                    'qty':      item['quantity'],
                    'min_qty':  item['min_qty'],
                    'cost':     item['unit_cost'],
                    'reason':   f'{mid} has {rul:.0f} days RUL, {fp:.0f}% failure risk',
                    'urgency':  'CRITICAL' if status == 'Out' else 'HIGH',
                })
        # Also flag any out-of-stock regardless
        if item['status'] == 'Out':
            if not any(a['inv_id'] == item['id'] for a in alerts):
                alerts.append({
                    'part':     item['part'],
                    'inv_id':   item['id'],
                    'machine':  ', '.join(item['compatible']),
                    'status':   'Out',
                    'qty':      0,
                    'min_qty':  item['min_qty'],
                    'cost':     item['unit_cost'],
                    'reason':   'Out of stock — reorder immediately',
                    'urgency':  'CRITICAL',
                })
    # Deduplicate by inv_id, keep highest urgency
    seen = {}
    for a in alerts:
        k = a['inv_id']
        if k not in seen or a['urgency'] == 'CRITICAL':
            seen[k] = a
    return list(seen.values())


def get_checklist(machine_id, failure_type):
    base = [
        {"id":1,"task":"Inspect all visible components",      "done":False},
        {"id":2,"task":"Check lubrication levels",            "done":False},
        {"id":3,"task":"Verify electrical connections",       "done":False},
        {"id":4,"task":"Test safety shutoff systems",         "done":False},
        {"id":5,"task":"Record sensor baseline readings",     "done":False},
        {"id":6,"task":"Clean filters and vents",             "done":False},
    ]
    extras = {
        "Overheating":    [{"id":7,"task":"Flush cooling system","done":False},
                           {"id":8,"task":"Check coolant level","done":False}],
        "Bearing Issue":  [{"id":7,"task":"Replace worn bearings","done":False},
                           {"id":8,"task":"Perform dynamic balancing","done":False}],
        "Valve Blockage": [{"id":7,"task":"Clean all valves","done":False},
                           {"id":8,"task":"Test pressure relief valve","done":False}],
        "Wear and Tear":  [{"id":7,"task":"Replace consumable parts","done":False},
                           {"id":8,"task":"Full mechanical overhaul","done":False}],
    }
    return base + extras.get(failure_type, [])

# ── SHIFT MANAGEMENT ──────────────────────────────────────────────────────────
SHIFTS = [
    {"shift":"Morning",   "time":"06:00 - 14:00", "operators":["Ravi Kumar","Priya Sharma"],  "supervisor":"Admin"},
    {"shift":"Afternoon", "time":"14:00 - 22:00", "operators":["Suresh Nair","Arun Patel"],   "supervisor":"manager1"},
    {"shift":"Night",     "time":"22:00 - 06:00", "operators":["Meena Raj","Vikram Singh"],   "supervisor":"engineer1"},
]

def get_current_shift():
    h = datetime.now().hour
    if   6  <= h < 14: return SHIFTS[0]
    elif 14 <= h < 22: return SHIFTS[1]
    else:              return SHIFTS[2]

# ── ENERGY CONSUMPTION ────────────────────────────────────────────────────────
def get_energy_data(machine_id):
    random.seed(hash(machine_id) % 300)
    base_kw = {"MCH-101":45,"MCH-102":75,"MCH-103":30,
               "MCH-104":90,"MCH-105":120,"MCH-106":200}.get(machine_id, 60)
    hours = list(range(0, 24))
    usage = [round(base_kw + random.uniform(-10, 15), 1) for _ in hours]
    total = round(sum(usage), 1)
    cost  = round(total * 8.5, 2)
    return {"hours": hours, "usage": usage,
            "total_kwh": total, "cost_inr": cost, "base_kw": base_kw}
